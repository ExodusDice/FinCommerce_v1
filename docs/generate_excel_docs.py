import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def parse_md_tables(md_path):
    with open(md_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()
        
    tables = []
    current_table = []
    
    for line in lines:
        stripped = line.strip()
        if stripped.startswith('|'):
            if stripped.startswith('|---') or stripped.startswith('| :---') or stripped.startswith('|:---'):
                continue
            cells = [c.strip() for c in stripped.split('|')[1:-1]]
            current_table.append(cells)
        else:
            if current_table:
                tables.append(current_table)
                current_table = []
                
    if current_table:
        tables.append(current_table)
        
    return tables

def create_styled_sheet(ws, table_data, title):
    ws.title = title
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid") # Indigo
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10, color="374151")
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    # Write Header
    headers = table_data[0]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    # Write Rows
    for row_idx, r_data in enumerate(table_data[1:], 2):
        fill = zebra_fill if row_idx % 2 == 1 else white_fill
        for col_idx, value in enumerate(r_data, 1):
            # Clean markdown link syntax
            clean_value = re.sub(r'\[([^\]]+)\]\([^\)]+\)', r'\1', value)
            cell = ws.cell(row=row_idx, column=col_idx, value=clean_value)
            cell.fill = fill
            cell.font = data_font
            cell.border = thin_border
            
            # Alignments
            if col_idx in [1, 3, 5, 7]: # IDs, sections, status
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
    # Auto-fit Column Widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45) # cap at 45 width

def compile_excel_rtm(md_path, xlsx_path):
    print(f"Compiling RTM to {xlsx_path}...")
    tables = parse_md_tables(md_path)
    if not tables:
        print("Error: No table found in RTM.md")
        return
        
    wb = Workbook()
    ws = wb.active
    create_styled_sheet(ws, tables[0], "Traceability Matrix")
    wb.save(xlsx_path)
    print(f"Generated RTM.xlsx successfully!")

def compile_excel_timeline(md_path, xlsx_path):
    print(f"Compiling Project Timeline to {xlsx_path}...")
    tables = parse_md_tables(md_path)
    if not tables or len(tables) < 2:
        print("Warning: Expected at least 2 tables in Project_Timeline.md")
        
    wb = Workbook()
    
    # Sheet 1: Milestones
    if len(tables) >= 1:
        ws1 = wb.active
        create_styled_sheet(ws1, tables[0], "Milestones & Gantt")
    
    # Sheet 2: RACI Matrix
    if len(tables) >= 2:
        ws2 = wb.create_sheet()
        create_styled_sheet(ws2, tables[1], "RACI Responsibility")
        
    wb.save(xlsx_path)
    print(f"Generated Project_Timeline.xlsx successfully!")

if __name__ == "__main__":
    docs_dir = os.path.dirname(__file__)
    
    rtm_md = os.path.join(docs_dir, "RTM.md")
    rtm_xlsx = os.path.join(docs_dir, "RTM.xlsx")
    if os.path.exists(rtm_md):
        compile_excel_rtm(rtm_md, rtm_xlsx)
        
    timeline_md = os.path.join(docs_dir, "Project_Timeline.md")
    timeline_xlsx = os.path.join(docs_dir, "Project_Timeline.xlsx")
    if os.path.exists(timeline_md):
        compile_excel_timeline(timeline_md, timeline_xlsx)
